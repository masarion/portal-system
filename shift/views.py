import json
from datetime import date, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .models import Manager, Project, ShiftSelection, ShiftType, Staff, Submission, normalize_name


def is_staff_user(user):
    return user.is_active and user.is_staff


# ── 案件一覧 ──────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_staff_user)
def project_list(request):
    projects = Project.objects.prefetch_related('staff_members', 'submissions')
    today = date.today()
    project_data = []
    for p in projects:
        total = p.staff_members.count()
        submitted = p.submissions.filter(submitted=True).count()
        project_data.append({
            'project': p,
            'total_staff': total,
            'submitted_count': submitted,
            'is_past_deadline': p.deadline < timezone.now(),
        })
    return render(request, 'shift/management/project_list.html', {
        'project_data': project_data,
        'today': today,
    })


# ── 案件作成・編集 ─────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_staff_user)
def project_create(request):
    if request.method == 'POST':
        return _save_project(request, None)
    # デフォルト：来月を初期値にする
    today = date.today()
    if today.month == 12:
        default_start = date(today.year + 1, 1, 1)
    else:
        default_start = date(today.year, today.month + 1, 1)
    import calendar
    last_day = calendar.monthrange(default_start.year, default_start.month)[1]
    default_end = date(default_start.year, default_start.month, last_day)
    return render(request, 'shift/management/project_form.html', {
        'page_title': '案件を作成',
        'default_start': default_start.isoformat(),
        'default_end': default_end.isoformat(),
        'shift_colors': SHIFT_COLORS,
    })


@login_required
@user_passes_test(is_staff_user)
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        return _save_project(request, project)
    shift_types = list(project.shift_types.values('id', 'name', 'short_name', 'time_range', 'color', 'order'))
    staff_list = list(project.staff_members.values('id', 'code', 'name'))
    managers = list(project.managers.values('id', 'role', 'name'))
    return render(request, 'shift/management/project_form.html', {
        'page_title': '案件を編集',
        'project': project,
        'shift_types_json': json.dumps(shift_types, ensure_ascii=False),
        'staff_json': json.dumps(staff_list, ensure_ascii=False),
        'managers_json': json.dumps(managers, ensure_ascii=False),
        'shift_colors': SHIFT_COLORS,
        'default_start': '',
        'default_end': '',
    })


def _save_project(request, project):
    name = request.POST.get('name', '').strip()
    deadline = request.POST.get('deadline', '').strip()
    start_date = request.POST.get('start_date', '').strip() or None
    end_date = request.POST.get('end_date', '').strip() or None
    auth_mode = request.POST.get('auth_mode', Project.AUTH_NAME)
    info_message = request.POST.get('info_message', '').strip()
    copy_guide_message = request.POST.get('copy_guide_message', '').strip()
    confirm_message = request.POST.get('confirm_message', '').strip()

    if not name or not deadline:
        messages.error(request, '案件名と提出期限は必須です。')
        return redirect('shift:project_create' if project is None else 'shift:project_edit', pk=project.pk if project else None)

    if project is None:
        project = Project()

    project.name = name
    project.auth_mode = auth_mode
    project.deadline = deadline
    project.start_date = start_date
    project.end_date = end_date
    project.info_message = info_message
    project.copy_guide_message = copy_guide_message
    project.confirm_message = confirm_message
    project.save()

    # シフト枠
    shift_data = json.loads(request.POST.get('shift_types_json', '[]'))
    project.shift_types.all().delete()
    for i, s in enumerate(shift_data):
        if s.get('name', '').strip():
            ShiftType.objects.create(
                project=project,
                name=s['name'].strip(),
                short_name=s.get('short_name', '').strip(),
                time_range=s.get('time_range', '').strip(),
                color=s.get('color', '#2563eb'),
                order=i,
            )

    # スタッフ
    staff_data = json.loads(request.POST.get('staff_json', '[]'))
    existing_staff_ids = set(project.staff_members.values_list('id', flat=True))
    posted_ids = set()
    for s in staff_data:
        sid = s.get('id')
        name_val = s.get('name', '').strip()
        if not name_val:
            continue
        if sid:
            try:
                st = project.staff_members.get(pk=sid)
                st.code = s.get('code', '').strip()
                st.name = name_val
                if auth_mode == Project.AUTH_CODE and s.get('password', '').strip():
                    st.password = s['password'].strip()
                st.save()
                posted_ids.add(st.pk)
                continue
            except Staff.DoesNotExist:
                pass
        st = Staff.objects.create(
            project=project,
            code=s.get('code', '').strip(),
            name=name_val,
            password=s.get('password', '').strip() if auth_mode == Project.AUTH_CODE else '',
        )
        posted_ids.add(st.pk)
    for sid in existing_staff_ids - posted_ids:
        Staff.objects.filter(pk=sid).delete()

    # 担当者
    managers_data = json.loads(request.POST.get('managers_json', '[]'))
    project.managers.all().delete()
    for m in managers_data:
        if m.get('name', '').strip():
            Manager.objects.create(project=project, role=m.get('role', '').strip(), name=m['name'].strip())

    messages.success(request, f'案件「{project.name}」を保存しました。')
    return redirect('shift:project_list')


@login_required
@user_passes_test(is_staff_user)
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        name = project.name
        project.delete()
        messages.success(request, f'案件「{name}」を削除しました。')
        return redirect('shift:project_list')
    return render(request, 'shift/management/project_confirm_delete.html', {'project': project})


# ── 提出状況ダッシュボード ───────────────────────────────────────────────────────

@login_required
@user_passes_test(is_staff_user)
def submission_dashboard(request, pk):
    project = get_object_or_404(Project, pk=pk)
    submissions = (
        project.submissions
        .select_related('staff')
        .prefetch_related('selections__shift_types')
        .order_by('staff__code', 'staff__name')
    )
    total = submissions.count()
    submitted_count = submissions.filter(submitted=True).count()
    return render(request, 'shift/management/submission_dashboard.html', {
        'project': project,
        'submissions': submissions,
        'total': total,
        'submitted_count': submitted_count,
        'unsubmitted_count': total - submitted_count,
        'is_past_deadline': project.deadline < timezone.now(),
    })


# ── 一覧表（スケジュール） ──────────────────────────────────────────────────────

@login_required
@user_passes_test(is_staff_user)
def submission_schedule(request, pk):
    project = get_object_or_404(Project, pk=pk)
    dates = project.get_dates()
    shift_types = project.shift_types.all()
    submissions = (
        project.submissions
        .filter(submitted=True)
        .select_related('staff')
        .prefetch_related('selections__shift_types')
        .order_by('staff__code', 'staff__name')
    )
    # {submission_id: {date: [shift_types]}}
    sel_map = {}
    for sub in submissions:
        sel_map[sub.pk] = {}
        for sel in sub.selections.all():
            sel_map[sub.pk][sel.date] = list(sel.shift_types.all())

    return render(request, 'shift/management/submission_schedule.html', {
        'project': project,
        'dates': dates,
        'shift_types': shift_types,
        'submissions': submissions,
        'sel_map': sel_map,
    })


# ── Excel出力 ─────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_staff_user)
def export_excel(request, pk):
    project = get_object_or_404(Project, pk=pk)
    dates = project.get_dates()
    submissions = (
        project.submissions
        .filter(submitted=True)
        .select_related('staff')
        .prefetch_related('selections__shift_types')
        .order_by('staff__code', 'staff__name')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = project.name[:31]

    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    DAY_NAMES = ['月', '火', '水', '木', '金', '土', '日']

    # ヘッダー行：氏名 + 日付
    ws.cell(row=1, column=1, value='氏名').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).border = border
    ws.column_dimensions['A'].width = 16

    for col_idx, d in enumerate(dates, start=2):
        day_name = DAY_NAMES[d.weekday()]
        cell = ws.cell(row=1, column=col_idx, value=f'{d.month}/{d.day}\n({day_name})')
        cell.font = header_font
        cell.fill = header_fill
        if d.weekday() == 5:
            cell.fill = PatternFill(fill_type='solid', fgColor='1d4ed8')
        elif d.weekday() == 6:
            cell.fill = PatternFill(fill_type='solid', fgColor='dc2626')
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 8
    ws.row_dimensions[1].height = 30

    # スタッフ行
    sel_map = {}
    for sub in submissions:
        sel_map[sub.pk] = {sel.date: [st.short_name or st.name[:3] for st in sel.shift_types.all()]
                           for sel in sub.selections.all()}

    for row_idx, sub in enumerate(submissions, start=2):
        name = sub.staff.name if sub.staff else 'スタッフ未登録'
        cell = ws.cell(row=row_idx, column=1, value=name)
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        if row_idx % 2 == 0:
            cell.fill = PatternFill(fill_type='solid', fgColor='F8FAFC')

        for col_idx, d in enumerate(dates, start=2):
            shifts = sel_map.get(sub.pk, {}).get(d, [])
            value = '\n'.join(shifts) if shifts else '―'
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.alignment = center
            c.border = border
            c.font = Font(size=9)
            if row_idx % 2 == 0:
                c.fill = PatternFill(fill_type='solid', fgColor='F8FAFC')

    ws.freeze_panes = 'B2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'シフト一覧_{project.name}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return response


# ── スタッフ提出画面（ログイン不要・案件ごとに1URL） ─────────────────────────────

def staff_submit(request, token):
    """案件トークンでアクセス→認証モードに応じて処理"""
    project = get_object_or_404(Project, submit_token=token)
    is_past_deadline = project.deadline < timezone.now()

    if project.auth_mode == Project.AUTH_CODE:
        # コード＋パスワードモード：認証画面へ
        return render(request, 'shift/staff_auth.html', {
            'project': project,
            'token': token,
            'is_past_deadline': is_past_deadline,
        })
    else:
        # 名前入力モード：名前入力＋カレンダーを同時表示
        # セッションから前回の名前を取得
        staff_name = request.session.get(f'staff_name_{project.pk}', '')
        submission = None
        existing = {}
        if staff_name:
            staff = project.staff_members.filter(name=staff_name).first()
            if staff:
                submission, _ = Submission.objects.get_or_create(project=project, staff=staff)
                for sel in submission.selections.prefetch_related('shift_types').all():
                    existing[sel.date.isoformat()] = [st.pk for st in sel.shift_types.all()]

        return render(request, 'shift/staff_submit.html', {
            'project': project,
            'token': token,
            'submission': submission,
            'staff_name': staff_name,
            'dates': project.get_dates(),
            'shift_types': project.shift_types.all(),
            'existing_json': json.dumps(existing),
            'is_past_deadline': is_past_deadline,
        })


def staff_auth(request, token):
    """コード＋パスワード認証処理（POST）"""
    project = get_object_or_404(Project, submit_token=token)
    if request.method != 'POST':
        return redirect('shift:staff_submit', token=token)

    code = request.POST.get('code', '').strip()
    password = request.POST.get('password', '').strip()

    staff = project.staff_members.filter(code=code, password=password).first()
    if not staff:
        is_past_deadline = project.deadline < timezone.now()
        return render(request, 'shift/staff_auth.html', {
            'project': project,
            'token': token,
            'error': 'コードまたはパスワードが正しくありません。',
            'is_past_deadline': is_past_deadline,
        })

    submission, _ = Submission.objects.get_or_create(project=project, staff=staff)
    existing = {}
    for sel in submission.selections.prefetch_related('shift_types').all():
        existing[sel.date.isoformat()] = [st.pk for st in sel.shift_types.all()]

    return render(request, 'shift/staff_submit.html', {
        'project': project,
        'token': token,
        'submission': submission,
        'staff_name': staff.name,
        'staff_code': staff.code,
        'dates': project.get_dates(),
        'shift_types': project.shift_types.all(),
        'existing_json': json.dumps(existing),
        'is_past_deadline': project.deadline < timezone.now(),
    })


@csrf_exempt
@require_http_methods(['POST'])
def staff_submit_save(request, token):
    project = get_object_or_404(Project, submit_token=token)

    if project.deadline < timezone.now():
        return JsonResponse({'success': False, 'message': '提出期限を過ぎています。'})

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'データ形式が無効です。'})

    staff_id = data.get('staff_id')
    staff_name = data.get('staff_name', '').strip()
    selections_data = data.get('selections', {})
    notes = data.get('notes', '').strip()

    # スタッフ特定
    if staff_id:
        staff = get_object_or_404(Staff, pk=staff_id, project=project)
    elif staff_name:
        normalized = normalize_name(staff_name)
        staff = project.staff_members.filter(name=normalized).first()
        if not staff:
            return JsonResponse({'success': False, 'message': f'「{staff_name}」はスタッフ一覧に見つかりません。名前を確認してください。'})
        # セッションに名前を保存
        request.session[f'staff_name_{project.pk}'] = normalized
    else:
        return JsonResponse({'success': False, 'message': 'スタッフ情報が不明です。'})

    submission, _ = Submission.objects.get_or_create(project=project, staff=staff)
    submission.selections.all().delete()
    for date_str, shift_ids in selections_data.items():
        try:
            d = date.fromisoformat(date_str)
        except ValueError:
            continue
        sel = ShiftSelection.objects.create(submission=submission, date=d)
        valid_ids = ShiftType.objects.filter(pk__in=shift_ids, project=project).values_list('pk', flat=True)
        sel.shift_types.set(valid_ids)

    submission.notes = notes
    submission.submitted = True
    submission.submitted_at = timezone.now()
    submission.save(update_fields=['notes', 'submitted', 'submitted_at', 'updated_at'])

    return JsonResponse({'success': True, 'message': '提出が完了しました。'})


# ── 定数 ─────────────────────────────────────────────────────────────────────

SHIFT_COLORS = [
    '#2563eb', '#16a34a', '#dc2626', '#d97706', '#7c3aed',
    '#0891b2', '#be185d', '#65a30d', '#ea580c', '#0369a1',
    '#9333ea', '#0d9488', '#b45309', '#4f46e5', '#059669',
]

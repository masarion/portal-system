import io
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.http import HttpResponse

from .models import AttendanceRecord, Employee, Workplace


def is_staff_user(user):
    return user.is_active and user.is_staff


# ──────────────── QRスキャン（認証不要） ────────────────

def scan_show(request, token):
    employee = get_object_or_404(Employee, qr_token=token, is_active=True)
    today    = timezone.localdate()
    record, _ = AttendanceRecord.objects.get_or_create(employee=employee, date=today)
    return render(request, 'attendance/scan.html', {'employee': employee, 'record': record})


def scan_store(request, token):
    if request.method != 'POST':
        return redirect('attendance:scan_show', token=token)
    employee = get_object_or_404(Employee, qr_token=token, is_active=True)
    today    = timezone.localdate()
    record, _ = AttendanceRecord.objects.get_or_create(employee=employee, date=today)
    now_time  = timezone.localtime().time()

    if not record.check_in:
        record.check_in = now_time
        record.save()
        msg = '出勤打刻が完了しました'
    elif not record.check_out:
        record.check_out = now_time
        record.save()
        msg = '退勤打刻が完了しました'
    else:
        msg = '本日の打刻は完了しています'

    messages.success(request, msg)
    return redirect('attendance:scan_show', token=token)


# ──────────────── 管理画面 ────────────────

@login_required
@user_passes_test(is_staff_user)
def dashboard(request):
    today      = timezone.localdate()
    total      = Employee.objects.filter(is_active=True).count()
    checked_in = AttendanceRecord.objects.filter(date=today, check_in__isnull=False, check_out__isnull=True).count()
    completed  = AttendanceRecord.objects.filter(date=today, check_in__isnull=False, check_out__isnull=False).count()
    not_yet    = total - checked_in - completed
    recent     = AttendanceRecord.objects.filter(date=today).select_related('employee__workplace').order_by('-updated_at')[:20]
    stats = [
        {'label': '全スタッフ', 'value': total,      'icon': 'people',      'color': 'primary'},
        {'label': '出勤中',     'value': checked_in,  'icon': 'person-check','color': 'success'},
        {'label': '退勤済み',   'value': completed,   'icon': 'check2-all',  'color': 'info'},
        {'label': '未出勤',     'value': not_yet,     'icon': 'person-dash', 'color': 'warning'},
    ]
    return render(request, 'attendance/management/dashboard.html', {
        'today': today, 'stats': stats, 'recent': recent,
    })


@login_required
@user_passes_test(is_staff_user)
def attendance_list(request):
    qs = AttendanceRecord.objects.select_related('employee__workplace').order_by('-date', 'employee')
    if request.GET.get('date_from'):
        qs = qs.filter(date__gte=request.GET['date_from'])
    if request.GET.get('date_to'):
        qs = qs.filter(date__lte=request.GET['date_to'])
    if request.GET.get('workplace_id'):
        qs = qs.filter(employee__workplace_id=request.GET['workplace_id'])
    status = request.GET.get('status')
    if status == 'completed':
        qs = qs.filter(check_in__isnull=False, check_out__isnull=False)
    elif status == 'working':
        qs = qs.filter(check_in__isnull=False, check_out__isnull=True)
    elif status == 'absent':
        qs = qs.filter(check_in__isnull=True)
    workplaces = Workplace.objects.all()
    return render(request, 'attendance/management/attendance_list.html', {
        'records': qs, 'workplaces': workplaces,
    })


@login_required
@user_passes_test(is_staff_user)
def attendance_export(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxlがインストールされていません')
        return redirect('attendance:attendance_list')

    qs = AttendanceRecord.objects.select_related('employee__workplace').order_by('-date', 'employee')
    if request.GET.get('date_from'):
        qs = qs.filter(date__gte=request.GET['date_from'])
    if request.GET.get('date_to'):
        qs = qs.filter(date__lte=request.GET['date_to'])
    if request.GET.get('workplace_id'):
        qs = qs.filter(employee__workplace_id=request.GET['workplace_id'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '勤怠一覧'
    headers = ['日付', '氏名', 'スタッフ番号', '所属', '出勤', '退勤', '勤務時間', '状態']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='2563EB')
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center')

    status_map = {'completed': '退勤済み', 'working': '出勤中', 'absent': '未出勤'}
    for r in qs:
        ws.append([
            str(r.date),
            r.employee.name,
            r.employee.employee_number,
            r.employee.workplace.name if r.employee.workplace else '-',
            str(r.check_in)[:5]  if r.check_in  else '-',
            str(r.check_out)[:5] if r.check_out else '-',
            r.work_duration,
            status_map.get(r.status, '-'),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'勤怠一覧_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ──────── スタッフ ────────

@login_required
@user_passes_test(is_staff_user)
def employee_list(request):
    qs = Employee.objects.select_related('workplace').order_by('employee_number')
    if request.GET.get('search'):
        q = request.GET['search']
        qs = qs.filter(name__icontains=q) | qs.filter(employee_number__icontains=q)
        qs = Employee.objects.filter(pk__in=qs.values('pk')).select_related('workplace').order_by('employee_number')
    if request.GET.get('workplace_id'):
        qs = qs.filter(workplace_id=request.GET['workplace_id'])
    if request.GET.get('is_active') in ('0', '1'):
        qs = qs.filter(is_active=request.GET['is_active'] == '1')
    workplaces = Workplace.objects.all()
    return render(request, 'attendance/management/employee_list.html', {'employees': qs, 'workplaces': workplaces})


@login_required
@user_passes_test(is_staff_user)
def employee_form(request, pk=None):
    employee   = get_object_or_404(Employee, pk=pk) if pk else None
    workplaces = Workplace.objects.all()
    errors     = {}
    if request.method == 'POST':
        name   = request.POST.get('name', '').strip()
        number = request.POST.get('employee_number', '').strip()
        wp_id  = request.POST.get('workplace_id') or None
        active = request.POST.get('is_active') == '1'
        if not name:
            errors['name'] = '氏名は必須です'
        if not number:
            errors['employee_number'] = 'スタッフ番号は必須です'
        if not errors:
            if employee:
                employee.name            = name
                employee.employee_number = number
                employee.workplace_id    = wp_id
                employee.is_active       = active
                employee.save()
                messages.success(request, 'スタッフ情報を更新しました')
            else:
                Employee.objects.create(name=name, employee_number=number, workplace_id=wp_id, is_active=active)
                messages.success(request, 'スタッフを登録しました')
            return redirect('attendance:employee_list')
    return render(request, 'attendance/management/employee_form.html', {
        'employee': employee, 'workplaces': workplaces, 'errors': errors,
    })


@login_required
@user_passes_test(is_staff_user)
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'スタッフを削除しました')
        return redirect('attendance:employee_list')
    return render(request, 'attendance/management/employee_confirm_delete.html', {'employee': employee})


@login_required
@user_passes_test(is_staff_user)
def employee_qr(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    scan_url = request.build_absolute_uri(f'/attendance/scan/{employee.qr_token}/')
    return render(request, 'attendance/management/employee_qr.html', {'employee': employee, 'scan_url': scan_url})


@login_required
@user_passes_test(is_staff_user)
def employee_import(request):
    if request.method == 'POST':
        text  = request.POST.get('csv_text', '')
        lines = re.split(r'\r\n|\r|\n', text.strip())
        added = 0
        for line in lines:
            line = line.strip()
            if not line:
                continue
            delim  = '\t' if '\t' in line else ','
            parts  = line.split(delim)
            number = parts[0].strip()
            name   = parts[1].strip() if len(parts) > 1 else number
            if not name or Employee.objects.filter(employee_number=number).exists():
                continue
            Employee.objects.create(name=name, employee_number=number)
            added += 1
        messages.success(request, f'{added}名をインポートしました')
        return redirect('attendance:employee_list')
    return redirect('attendance:employee_list')


# ──────── 所属場所 ────────

@login_required
@user_passes_test(is_staff_user)
def workplace_list(request):
    workplaces = Workplace.objects.annotate_employee_count() if hasattr(Workplace.objects, 'annotate_employee_count') else Workplace.objects.all()
    from django.db.models import Count
    workplaces = Workplace.objects.annotate(employee_count=Count('employees')).order_by('order', 'name')
    return render(request, 'attendance/management/workplace_list.html', {'workplaces': workplaces})


@login_required
@user_passes_test(is_staff_user)
def workplace_form(request, pk=None):
    workplace = get_object_or_404(Workplace, pk=pk) if pk else None
    errors    = {}
    if request.method == 'POST':
        name  = request.POST.get('name', '').strip()
        order = int(request.POST.get('order', 0) or 0)
        if not name:
            errors['name'] = '場所名は必須です'
        if not errors:
            if workplace:
                workplace.name  = name
                workplace.order = order
                workplace.save()
                messages.success(request, '所属場所を更新しました')
            else:
                Workplace.objects.create(name=name, order=order)
                messages.success(request, '所属場所を登録しました')
            return redirect('attendance:workplace_list')
    return render(request, 'attendance/management/workplace_form.html', {'workplace': workplace, 'errors': errors})


@login_required
@user_passes_test(is_staff_user)
def workplace_delete(request, pk):
    workplace = get_object_or_404(Workplace, pk=pk)
    if request.method == 'POST':
        workplace.delete()
        messages.success(request, '所属場所を削除しました')
        return redirect('attendance:workplace_list')
    return render(request, 'attendance/management/workplace_confirm_delete.html', {'workplace': workplace})
